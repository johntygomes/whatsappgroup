import time
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains

import openpyxl
import xlrd

driver=webdriver.Chrome()
driver.get("https://web.whatsapp.com/")
action = ActionChains(driver)
time.sleep(25) 

def extractnumbers(sheet,row,groupname,xlfile):
    addrows=row
    thecrntgrup=groupname
    file=xlfile
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[sheet]
                        
    grup=driver.find_element_by_xpath('/html/body/div[1]/div/div/div[3]/div/div[2]/div[1]/div/div/div[1]/div/div ')
    grup.click()
    time.sleep(2)

    numberslist=driver.find_element_by_xpath('/html/body/div[1]/div/div/div[4]/div/header/div[2]/div[2]/span')
    numbersarray=numberslist.text.split(",")

    for i in range(0,len(numbersarray)):
        ws.cell(row=i+addrows, column=1, value=numbersarray[i])
        ws.cell(row=i+addrows, column=2, value=thecrntgrup)    
    wb.save(file)


